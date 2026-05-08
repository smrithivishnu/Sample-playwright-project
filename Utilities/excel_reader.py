import random
import os

# Try to import openpyxl at module level with proper error handling
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Will use fallback data.")


def read_port_list_excel():
    """
    Common function to read PortList.xlsx file
    Returns list of dictionaries containing port data
    """
    if OPENPYXL_AVAILABLE:
        try:
            # Try to find PortList.xlsx in project root
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(current_dir)
            excel_file = os.path.join(project_root, "PortList.xlsx")
            
            # Verify file exists
            if not os.path.exists(excel_file):
                print(f"PortList.xlsx not found at {excel_file}. Using fallback data.")
                return get_fallback_ports()
            
            # Read Excel file using openpyxl
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            if not sheet:
                print("No active sheet found in Excel file. Using fallback data.")
                workbook.close()
                return get_fallback_ports()
            
            # Read data rows
            ports = []
            max_row = sheet.max_row
            print(f"Reading ports from Excel file: {excel_file} (rows: {max_row})")
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Ensure row has enough columns and required data exists
                if not row or len(row) < 3:  # Need at least 3 columns for port name, code, country
                    continue
                
                port_name = row[1]  # Column B - Port Name
                port_code = row[2]  # Column C - Port Code  
                port_country = row[3] if len(row) > 3 else ""  # Column D - Country
                
                # Skip rows with missing critical data
                if not port_code or not port_name:
                    continue
                
                # Skip empty strings after stripping
                port_name = str(port_name).strip() if port_name else ""
                port_code = str(port_code).strip() if port_code else ""
                port_country = str(port_country).strip() if port_country else ""
                
                if not port_code or not port_name:
                    continue
                
                port_dict = {
                    "VIP_EXT_REF": port_code,
                    "VIP_PORT_NAME": port_name,
                    "VIP_PORT_COUNTRY": port_country
                }
                ports.append(port_dict)
            
            workbook.close()
            print(f"Successfully read {len(ports)} ports from Excel")
            return ports if ports else get_fallback_ports()
            
        except FileNotFoundError:
            print(f"PortList.xlsx file not found. Using fallback data.")
            return get_fallback_ports()
        except Exception as e:
            print(f"Error reading Excel file: {str(e)}. Using fallback data.")
            return get_fallback_ports()
    else:
        # Use fallback data if openpyxl is not available
        print("Using fallback port data (openpyxl not available)")
        return get_fallback_ports()


def get_fallback_ports():
    """
    Fallback port data when Excel reading fails
    Returns list of dictionaries containing port data
    """
    return [
        {"VIP_EXT_REF": "AMS", "VIP_PORT_NAME": "AMSTERDAM", "VIP_PORT_COUNTRY": "NETHERLANDS"},
        {"VIP_EXT_REF": "ROT", "VIP_PORT_NAME": "ROTTERDAM", "VIP_PORT_COUNTRY": "NETHERLANDS"},
        {"VIP_EXT_REF": "HAM", "VIP_PORT_NAME": "HAMBURG", "VIP_PORT_COUNTRY": "GERMANY"},
        {"VIP_EXT_REF": "LON", "VIP_PORT_NAME": "LONDON", "VIP_PORT_COUNTRY": "UNITED KINGDOM"},
        {"VIP_EXT_REF": "PAR", "VIP_PORT_NAME": "PARIS", "VIP_PORT_COUNTRY": "FRANCE"},
        {"VIP_EXT_REF": "BAR", "VIP_PORT_NAME": "BARCELONA", "VIP_PORT_COUNTRY": "SPAIN"},
        {"VIP_EXT_REF": "LIS", "VIP_PORT_NAME": "LISBON", "VIP_PORT_COUNTRY": "PORTUGAL"},
        {"VIP_EXT_REF": "ROM", "VIP_PORT_NAME": "ROME", "VIP_PORT_COUNTRY": "ITALY"},
        {"VIP_EXT_REF": "ATH", "VIP_PORT_NAME": "ATHENS", "VIP_PORT_COUNTRY": "GREECE"},
        {"VIP_EXT_REF": "IST", "VIP_PORT_NAME": "ISTANBUL", "VIP_PORT_COUNTRY": "TURKEY"},
        {"VIP_EXT_REF": "CPT", "VIP_PORT_NAME": "CAPE TOWN", "VIP_PORT_COUNTRY": "SOUTH AFRICA"},
        {"VIP_EXT_REF": "MUM", "VIP_PORT_NAME": "MUMBAI", "VIP_PORT_COUNTRY": "INDIA"},
        {"VIP_EXT_REF": "SIN", "VIP_PORT_NAME": "SINGAPORE", "VIP_PORT_COUNTRY": "SINGAPORE"},
        {"VIP_EXT_REF": "TOK", "VIP_PORT_NAME": "TOKYO", "VIP_PORT_COUNTRY": "JAPAN"},
        {"VIP_EXT_REF": "SYD", "VIP_PORT_NAME": "SYDNEY", "VIP_PORT_COUNTRY": "AUSTRALIA"},
        {"VIP_EXT_REF": "NYC", "VIP_PORT_NAME": "NEW YORK", "VIP_PORT_COUNTRY": "UNITED STATES"},
        {"VIP_EXT_REF": "LAX", "VIP_PORT_NAME": "LOS ANGELES", "VIP_PORT_COUNTRY": "UNITED STATES"},
        {"VIP_EXT_REF": "CHI", "VIP_PORT_NAME": "CHICAGO", "VIP_PORT_COUNTRY": "UNITED STATES"},
        {"VIP_EXT_REF": "MIA", "VIP_PORT_NAME": "MIAMI", "VIP_PORT_COUNTRY": "UNITED STATES"},
        {"VIP_EXT_REF": "DUB", "VIP_PORT_NAME": "DUBAI", "VIP_PORT_COUNTRY": "UAE"}
    ]


def get_random_port():
    """
    Get one random port from PortList.xlsx or fallback data
    Returns dictionary with port data
    """
    import time
    
    all_ports = read_port_list_excel()
    
    if not all_ports:
        # Return default port data if no ports available
        return {
            "VIP_EXT_REF": "AMS",
            "VIP_PORT_NAME": "AMSTERDAM", 
            "VIP_PORT_COUNTRY": "NETHERLANDS"
        }
    
    # Use time-based seed for better randomness
    current_time = int(time.time() * 1000)
    random.seed(current_time)
    
    # Randomly select one port
    selected_port = random.choice(all_ports)
    
    return {
        "VIP_EXT_REF": selected_port.get("VIP_EXT_REF", ""),
        "VIP_PORT_NAME": selected_port.get("VIP_PORT_NAME", ""),
        "VIP_PORT_COUNTRY": selected_port.get("VIP_PORT_COUNTRY", "")
    }
